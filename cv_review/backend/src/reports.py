"""High-level candidate reports for decision makers (Excel & PDF)."""
import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

AREAS_ORDER = ["Infrastructure", "Networking", "Platform", "Data", "Other"]


def _escape_paragraph_text(s: str, max_chars: int = 2000) -> str:
    """Escape for ReportLab Paragraph and convert newlines to <br/>. Keep more text for PDF."""
    if not s or not isinstance(s, str):
        return "—"
    s = s[:max_chars].strip()
    s = s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    s = re.sub(r"\r\n|\r|\n", "<br/>", s)
    return s or "—"


def _record_to_row(record: Dict[str, Any]) -> Dict[str, Any]:
    """Extract high-level fields from a stored analysis record."""
    result = record.get("result") or {}
    area_scores = result.get("area_scores") or {}
    best_specs = result.get("best_specializations") or []
    education_list = result.get("education_list") or []
    education_str = "; ".join(str(x).strip() for x in education_list if x)[:200]
    return {
        "filename": str(record.get("filename") or "—"),
        "timestamp": str(record.get("timestamp") or ""),
        "most_fitted_area": str(result.get("most_fitted_area") or "—"),
        "recommended_role": str(result.get("recommended_role") or "—"),
        "candidate_summary": (str(result.get("candidate_summary") or "—"))[:500],
        "recommendation_reason": (str(result.get("recommendation_reason") or "—"))[:300],
        "area_scores": area_scores,
        "top_specializations": ", ".join(str(s.get("specialization") or "") for s in best_specs[:5]),
        "education": education_str or "—",
        "analysis_time_seconds": record.get("analysis_time_seconds", 0) or 0,
    }


def build_excel_report(records: List[Dict[str, Any]]) -> bytes:
    """Build an Excel workbook with one summary sheet (high-level for decision makers)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate Summary"

    # Header
    headers = [
        "Candidate",
        "Date",
        "Best-fit area",
        "Recommended role",
        "Summary",
        "Recommendation reason",
        "Infra",
        "Network",
        "Platform",
        "Data",
        "Other",
        "Top specializations",
        "Education",
        "AI time (s)",
    ]
    header_fill = PatternFill(start_color="4285F4", end_color="4285F4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, record in enumerate(records, 2):
        row_data = _record_to_row(record)
        area_scores = row_data["area_scores"]
        ws.cell(row=row_idx, column=1, value=row_data["filename"])
        ws.cell(row=row_idx, column=2, value=row_data["timestamp"][:10] if row_data["timestamp"] else "—")
        ws.cell(row=row_idx, column=3, value=row_data["most_fitted_area"])
        ws.cell(row=row_idx, column=4, value=row_data["recommended_role"])
        ws.cell(row=row_idx, column=5, value=row_data["candidate_summary"])
        ws.cell(row=row_idx, column=6, value=row_data["recommendation_reason"])
        for i, area in enumerate(AREAS_ORDER):
            ws.cell(row=row_idx, column=7 + i, value=area_scores.get(area, "—"))
        ws.cell(row=row_idx, column=12, value=row_data["top_specializations"])
        ws.cell(row=row_idx, column=13, value=row_data["education"])
        ws.cell(row=row_idx, column=14, value=row_data["analysis_time_seconds"])
        for c in range(1, 15):
            ws.cell(row=row_idx, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 35
    for col in "GHIJ":
        ws.column_dimensions[col].width = 8
    ws.column_dimensions["K"].width = 8
    ws.column_dimensions["L"].width = 32
    ws.column_dimensions["M"].width = 28
    ws.column_dimensions["N"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_report(records: List[Dict[str, Any]]) -> bytes:
    """Build a PDF with one section per candidate (high-level for decision makers)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=inch,
        leftMargin=inch,
        topMargin=inch,
        bottomMargin=inch,
    )
    styles = getSampleStyleSheet()
    # Style for table cell paragraphs: wrap text, smaller font
    from reportlab.lib.styles import ParagraphStyle
    cell_style = ParagraphStyle(
        name="CellWrap",
        parent=styles["Normal"],
        fontSize=9,
        leading=11,
        wordWrap="CJK",  # enables word wrap
    )
    story = []

    title = Paragraph(
        "<b>CV Review — Candidate Report (Google Team)</b><br/><font size=9 color=gray>High-level summary for decision making</font>",
        styles["Title"],
    )
    story.append(title)
    story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    story.append(Spacer(1, 0.3 * inch))

    for record in records:
        row_data = _record_to_row(record)
        result = record.get("result") or {}
        filename = _escape_paragraph_text(str(record.get("filename") or "—"), 200)
        # Full summary and reason for PDF (with wrapping via Paragraph)
        summary_full = (result.get("candidate_summary") or "—")
        reason_full = (result.get("recommendation_reason") or "—")
        summary_para = Paragraph(_escape_paragraph_text(summary_full, 2000), cell_style)
        reason_para = Paragraph(_escape_paragraph_text(reason_full, 1000), cell_style)

        story.append(Paragraph(f"<b>Candidate: {filename}</b>", styles["Heading2"]))
        story.append(Spacer(1, 0.1 * inch))

        # Key decision table: use Paragraph in cells so text wraps and shows in full
        data = [
            ["Best-fit area", row_data["most_fitted_area"]],
            ["Recommended role", row_data["recommended_role"]],
            ["Summary", summary_para],
            ["Recommendation reason", reason_para],
        ]
        t = Table(data, colWidths=[1.8 * inch, 4.2 * inch])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8F0FE")),
                    ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1967D2")),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (0, -1), 9),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(t)
        story.append(Spacer(1, 0.15 * inch))

        # Area scores
        area_vals = [str(row_data["area_scores"].get(a, "—")) for a in AREAS_ORDER]
        score_headers = ["Infrastructure", "Networking", "Platform", "Data", "Other"]
        score_table = Table([score_headers, area_vals], colWidths=[1.1 * inch] * 5)
        score_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F1F3F4")),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]
            )
        )
        story.append(Paragraph("<b>Area scores (1–5)</b>", styles["Normal"]))
        story.append(score_table)
        story.append(Spacer(1, 0.1 * inch))
        top_spec = _escape_paragraph_text(row_data.get("top_specializations") or "—", 300)
        education = _escape_paragraph_text(row_data.get("education") or "—", 400)
        story.append(Paragraph(f"<b>Top specializations:</b> {top_spec}", styles["Normal"]))
        story.append(Paragraph(f"<b>Education:</b> {education}", styles["Normal"]))
        story.append(Spacer(1, 0.35 * inch))

    doc.build(story)
    return buf.getvalue()


def get_records_for_report(
    storage,
    analysis_ids: Optional[List[str]] = None,
    area_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 200,
) -> List[Dict[str, Any]]:
    """Fetch analysis records from storage, optionally filtered by ids, area, or date range.
    date_from / date_to: ISO date strings YYYY-MM-DD (inclusive)."""
    analyses = storage._read()
    # Newest first
    analyses = sorted(analyses, key=lambda x: x.get("timestamp", ""), reverse=True)[: limit * 2]

    if analysis_ids:
        id_set = set(analysis_ids)
        analyses = [a for a in analyses if a.get("id") in id_set]
    if area_filter:
        analyses = [a for a in analyses if (a.get("result") or {}).get("most_fitted_area") == area_filter]

    if date_from or date_to:
        def _date_ok(ts: str) -> bool:
            if not ts:
                return False
            # timestamp is ISO like 2025-02-23T20:00:00.000Z; compare date part
            record_date = (ts.split("T")[0] if "T" in ts else ts[:10])[:10]
            if date_from and record_date < date_from:
                return False
            if date_to and record_date > date_to:
                return False
            return True

        analyses = [a for a in analyses if _date_ok(a.get("timestamp") or "")]

    return analyses[:limit]
